from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from django.utils import timezone

from .models import Inventory, ActivityLog, ExcelFile, BackupFile
import openpyxl
import os
import shutil
import datetime

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            ActivityLog.objects.create(
                user=user,
                action='login',
                details=f"Login berhasil dari {request.META.get('REMOTE_ADDR')}"
            )
            return redirect('dashboard')
        else:
            messages.error(request, 'Username atau password salah')
    
    return render(request, 'dashboard/login.html')

@login_required
def dashboard(request):
    inventory_items = Inventory.objects.all()
    context = {
        'inventory_items': inventory_items,
        'active_menu': 'dashboard'
    }
    return render(request, 'dashboard/dashboard.html', context)

@login_required
def upload_file(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Save the uploaded file
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'uploads'))
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)
        
        # Save file record
        excel_record = ExcelFile.objects.create(
            file=f'uploads/{filename}',
            uploaded_by=request.user,
            filename=filename
        )
        
        # Process Excel file
        try:
            # Clear existing inventory
            Inventory.objects.all().delete()
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            # Skip header row
            rows = list(sheet.rows)[1:]
            
            # Process rows
            for row in rows:
                kode = row[0].value
                nama_barang = row[1].value
                kategori = row[2].value
                harga_jual = row[3].value or 0
                total_stok = row[4].value or 0
                
                if kode and nama_barang:  # Ensure required fields exist
                    Inventory.objects.create(
                        kode=kode,
                        nama_barang=nama_barang,
                        kategori=kategori or '',
                        harga_jual=harga_jual,
                        total_stok=total_stok
                    )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action='upload',
                details=f"Upload file Excel: {filename}"
            )
            
            messages.success(request, f'File {filename} berhasil diupload dan diproses')
        except Exception as e:
            messages.error(request, f'Error memproses file: {str(e)}')
            ActivityLog.objects.create(
                user=request.user,
                action='upload',
                details=f"Error upload file Excel: {filename}. Error: {str(e)}"
            )
    
    context = {
        'active_menu': 'upload'
    }
    return render(request, 'dashboard/upload.html', context)

@login_required
def backup_file(request):
    backups = BackupFile.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        # Create backup of current inventory
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Barang Semua Cabang"
            
            # Add headers
            headers = ['Kode', 'Nama Barang', 'Kategori', 'Harga Jual', 'Total Stok']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
            
            # Add data
            inventory_items = Inventory.objects.all()
            for row_num, item in enumerate(inventory_items, 2):
                sheet.cell(row=row_num, column=1, value=item.kode)
                sheet.cell(row=row_num, column=2, value=item.nama_barang)
                sheet.cell(row=row_num, column=3, value=item.kategori)
                sheet.cell(row=row_num, column=4, value=float(item.harga_jual))
                sheet.cell(row=row_num, column=5, value=item.total_stok)
            
            # Save the file
            timestamp = timezone.now().strftime('%Y_%m_%d_%H_%M_%S')
            filename = f'Backup_Inventory_{timestamp}.xlsx'
            backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            file_path = os.path.join(backup_dir, filename)
            wb.save(file_path)
            
            # Create backup record
            backup = BackupFile.objects.create(
                file=f'backups/{filename}',
                created_by=request.user,
                filename=filename
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action='backup',
                details=f"Backup dibuat: {filename}"
            )
            
            messages.success(request, f'Backup berhasil dibuat: {filename}')
        except Exception as e:
            messages.error(request, f'Error membuat backup: {str(e)}')
            ActivityLog.objects.create(
                user=request.user,
                action='backup',
                details=f"Error membuat backup. Error: {str(e)}"
            )
    
    context = {
        'backups': backups,
        'active_menu': 'backup'
    }
    return render(request, 'dashboard/backup.html', context)

@login_required
def activity_logs(request):
    logs = ActivityLog.objects.all()
    context = {
        'logs': logs,
        'active_menu': 'logs'
    }
    return render(request, 'dashboard/logs.html', context)

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action='password',
                details="Password berhasil diubah"
            )
            
            messages.success(request, 'Password berhasil diubah')
            return redirect('change_password')
        else:
            messages.error(request, 'Mohon perbaiki error di bawah')
    else:
        form = PasswordChangeForm(request.user)
    
    context = {
        'form': form,
        'active_menu': 'password'
    }
    return render(request, 'dashboard/change_password.html', context)

@login_required
def logout_view(request):
    # Log activity before logout
    ActivityLog.objects.create(
        user=request.user,
        action='logout',
        details=f"Logout dari {request.META.get('REMOTE_ADDR')}"
    )
    
    logout(request)
    return redirect('login')
